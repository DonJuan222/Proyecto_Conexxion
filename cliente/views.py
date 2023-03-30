from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView
# para redirigir a otras paginas
from django.http import HttpResponseRedirect, HttpResponse
# Mensajes de formulario
from django.contrib import messages

# Importando los modelos
from cliente.models import *
from core.models import *

from core.funciones import *

from django.views import View
from cliente.forms import *
# Create your views here.

# formularios dinamicos
from django.forms import formset_factory

#Decoradores permisos y Login
from django.contrib.auth.decorators import login_required

#Libreria para hacer los reportes o trabajar con Excel
from openpyxl import Workbook


from django.db.models import OuterRef, Subquery

#Librerias para trabajar con Imagenes
from django.shortcuts import render
from django.http import HttpResponse
from PIL import Image, ImageDraw, ImageFont
from io import BytesIO

#Libreria para trabajar con Tiempo
from datetime import date 

# Crea una lista de los clientes, 10 por pagina------------------------------------------------------
class ListarClientes(LoginRequiredMixin, View):
    login_url = '/login'
    redirect_field_name = None

    def get(self, request):
        # Obtener los clientes con la fecha de pago y fecha de vencimiento más recientes de su factura
        # ultima_factura = Factura.objects.filter(cliente=OuterRef('pk')).order_by('-fecha_pago', '-fecha_vencimiento')
        
        ultima_factura = Factura.objects.filter(cliente=OuterRef('pk'), tipo_pago='#Recibo').order_by('-fecha_pago', '-fecha_vencimiento')
        clientes = Cliente.objects.annotate(fecha_pago=Subquery(ultima_factura.values('fecha_pago')[:1]),
                                             fecha_vencimiento=Subquery(ultima_factura.values('fecha_vencimiento')[:1]))

        contexto = {'tabla': clientes}
        contexto = complementarContexto(contexto, request.user)

        return render(request, 'cliente/listarClientes.html', contexto)
    
    def post(self, request, cliente_id):
        cliente = Cliente.objects.get(pk=cliente_id)
        estado_nuevo = request.POST['estado']
        cliente.estado = estado_nuevo
        cliente.save()
        return redirect('listarClientes')
    
# Fin de vista---------------------------------------------------------------------------------------

# Crea una lista de los clientes, 10 por pagina------------------------------------------------------
class ListarClientesRetirados(LoginRequiredMixin, View):
    login_url = '/login'
    redirect_field_name = None

    def get(self, request):
        from django.db import models
        # Saca una lista de todos los clientes de la BDD
        clientes_retirados = ClienteRetirado.objects.all()
        contexto = {'tabla': clientes_retirados}
        contexto = complementarContexto(contexto, request.user)

        return render(request, 'cliente/retiroCliente.html', contexto)
# Fin de vista---------------------------------------------------------------------------------------


# Crea y procesa un formulario para agregar a un cliente---------------------------------------------
class AgregarCliente(LoginRequiredMixin, View):
    login_url = '/login'
    redirect_field_name = None

    def post(self, request):
        # Crea una instancia del formulario y la llena con los datos:
        form = ClienteFormulario(request.POST)
        # Revisa si es valido:

        if form.is_valid():
            # Procesa y asigna los datos con form.cleaned_data como se requiere
            ip = form.cleaned_data['ip']
            cedula = form.cleaned_data['cedula']
            nombre = form.cleaned_data['nombre']
            apellido = form.cleaned_data['apellido']
            telefono_uno = form.cleaned_data['telefono_uno']
            telefonos_dos = form.cleaned_data['telefonos_dos']
            valor_instalacion = form.cleaned_data['valor_instalacion']
            fecha_instalacion = form.cleaned_data['fecha_instalacion']
            direccion = form.cleaned_data['direccion']
            estado = form.cleaned_data['estado']
            descripcion = form.cleaned_data['descripcion']
            municipio = form.cleaned_data['municipio']
            equipos = form.cleaned_data['equipos']         
            tipo_instalacion = form.cleaned_data['tipo_instalacion']  
            cap_megas = form.cleaned_data['cap_megas']        
            ap = form.cleaned_data['ap']           

            cliente = Cliente(ip=ip, cedula=cedula, nombre=nombre, apellido=apellido, telefono_uno=telefono_uno,
                              telefonos_dos=telefonos_dos, valor_instalacion=valor_instalacion, fecha_instalacion=fecha_instalacion,
                              direccion=direccion,estado=estado,municipio=municipio, tipo_instalacion=tipo_instalacion,
                              descripcion=descripcion,equipos=equipos,cap_megas=cap_megas,ap=ap)
            cliente.save()
            form = ClienteFormulario()

            messages.success(
                request, 'Ingresado exitosamente bajo la ID %s.' % cliente.id)
            request.session['clienteProcesado'] = 'agregado'
            return HttpResponseRedirect("/listarClientes")
        else:
            # De lo contrario lanzara el mismo formulario

            return render(request, 'cliente/agregarCliente.html', {'form': form})

    def get(self, request):

        form = ClienteFormulario()
        # Envia al usuario el formulario para que lo llene
        contexto = {'form': form,
                    'modo': request.session.get('clienteProcesado')}
        contexto = complementarContexto(contexto, request.user)
        return render(request, 'cliente/agregarCliente.html', contexto)
# Fin de vista---------------------------------------------------------------------------------------


# Muestra el mismo formulario del cliente pero con los datos a editar--------------------------------
class EditarCliente(LoginRequiredMixin, View):
    login_url = '/login'
    redirect_field_name = None

    def post(self, request, p):
        # Crea una instancia del formulario y la llena con los datos:
        cliente = Cliente.objects.get(id=p)
        form = ClienteFormulario(request.POST, instance=cliente)
        # Revisa si es valido:

        if form.is_valid():
            # Procesa y asigna los datos con form.cleaned_data como se requiere
            ip = form.cleaned_data['ip']
            cedula = form.cleaned_data['cedula']
            nombre = form.cleaned_data['nombre']
            apellido = form.cleaned_data['apellido']
            telefono_uno = form.cleaned_data['telefono_uno']
            telefonos_dos = form.cleaned_data['telefonos_dos']
            valor_instalacion = form.cleaned_data['valor_instalacion']
            fecha_instalacion = form.cleaned_data['fecha_instalacion']
            direccion = form.cleaned_data['direccion']
            estado = form.cleaned_data['estado']
            descripcion = form.cleaned_data['descripcion']
            municipio = form.cleaned_data['municipio']
            equipos = form.cleaned_data['equipos']         
            tipo_instalacion = form.cleaned_data['tipo_instalacion']  
            cap_megas = form.cleaned_data['cap_megas']  
            ap = form.cleaned_data['ap']  

            cliente.ip = ip
            cliente.cedula = cedula
            cliente.nombre = nombre
            cliente.apellido = apellido
            cliente.telefono_uno = telefono_uno
            cliente.telefonos_dos = telefonos_dos
            cliente.valor_instalacion = valor_instalacion
            cliente.fecha_instalacion = fecha_instalacion
            cliente.direccion = direccion
            cliente.estado = estado
            cliente.descripcion = descripcion
            cliente.municipio = municipio
            cliente.equipos = equipos
            cliente.tipo_instalacion = tipo_instalacion
            cliente.cap_megas = cap_megas
            cliente.ap = ap
            
            cliente.save()
            form = ClienteFormulario(instance=cliente)

            messages.success(
                request, 'Actualizado exitosamente el cliente de ID %s.' % p)
            request.session['clienteProcesado'] = 'editado'
            return HttpResponseRedirect("/listarClientes" )
        else:
            # De lo contrario lanzara el mismo formulario
            messages.success(request, 'La Ip o la Cedula ya estan registradas')
            return HttpResponseRedirect("/editarCliente/%s" % cliente.id)

    def get(self, request, p):
        cliente = Cliente.objects.get(id=p)
        form = ClienteFormulario(instance=cliente)
        # Envia al usuario el formulario para que lo llene
        contexto = {'form': form, 'modo': request.session.get(
            'clienteProcesado'), 'editar': True}
        contexto = complementarContexto(contexto, request.user)
        return render(request, 'cliente/agregarCliente.html', contexto)
# Fin de vista---------------------------------------------------------------------------------------


# # Funcion que permite importar un archivo csv y asi alimentar la base de datos---------------------
# def import_clientes(request):
#     if request.method == "POST":
#         csv_file = request.FILES['file']
#         if not csv_file.name.endswith('.csv'):
#             messages.error(request, 'El archivo no es un .csv')
#         data_set = csv_file.read().decode('UTF-8')
#         io_string = io.StringIO(data_set)
#         next(io_string)
#         df = pd.read_csv(io_string, delimiter=",")
#         for index, row in df.iterrows():
#             _, created = Cliente.objects.update_or_create(
#                 ip=row['ip'],
#                 defaults={
#                     'cedula': row['cedula'],
#                     'nombre': row['nombre'],
#                     'apellido': row['apellido'],
#                     'telefono_uno': row['telefono_uno'],
#                     'telefonos_dos': row['telefonos_dos'],
#                     'mensualidad': row['mensualidad'],
#                     'fecha_instalacion': row['fecha_instalacion'],
#                     'direccion': row['direccion'],
#                     'vereda': row['vereda'],
#                     'tipo_instalacion': row['tipo_instalacion'],
#                     'status': row['status'],
#                     'descripcion': row['descripcion'],
#                     'id_Municipio': row['id_Municipio'],
#                     'id_Pueblo': row['id_Pueblo'],
#                     'id_Estado': row['id_Estado']
#                 }
#             )
#         messages.success(request, 'Archivo importado exitosamente!')
#         return redirect("listarClientes")
#     return render(request, "cliente/importarClientes.html")
# # Fin de vista-------------------------------------------------------------------------------------


# Elimina usuarios,clientes -------------------------------------------------------------------------
class Eliminar(LoginRequiredMixin, View):
    login_url = '/login'
    redirect_field_name = None
 
    def get(self, request, modo, p):
        if modo == 'cliente':
            cliente = Cliente.objects.get(id=p)
            # Cambiar el estado del cliente a retirado
            cliente.estado = 'Retiros'
            cliente.save()
            # Mover el cliente a ClienteRetirado
            cliente_retirado = ClienteRetirado(
                ip=cliente.ip,
                cedula=cliente.cedula,
                nombre=cliente.nombre,
                apellido=cliente.apellido,
                telefono_uno=cliente.telefono_uno,
                telefonos_dos=cliente.telefonos_dos,
                valor_instalacion=cliente.valor_instalacion,
                fecha_instalacion=cliente.fecha_instalacion,
                direccion=cliente.direccion,
                estado=cliente.estado,
                descripcion=cliente.descripcion,
                municipio=cliente.municipio,
                equipos=cliente.equipos,
                tipo_instalacion=cliente.tipo_instalacion,
                cap_megas=cliente.cap_megas,
                ap=cliente.ap
            )
            cliente_retirado.save()

            # Mover las facturas del cliente a FacturaRetirada
            facturas = Factura.objects.filter(cliente=cliente)
            for factura in facturas:
                factura_retirada = FacturaRetirada(
                    cliente_retirado=cliente_retirado,
                    detalle=factura.detalle,
                    tipo_pago=factura.tipo_pago,
                    valor_pago=factura.valor_pago,
                    fecha_pago=factura.fecha_pago,
                    fecha_vencimiento=factura.fecha_vencimiento
                )
                factura_retirada.save()

            # Eliminar el cliente original y sus facturas
            facturas.delete()
            cliente.delete()

            messages.success(
                request, 'Cliente de ID %s movido a clientes retirados exitosamente.' % p)
            return HttpResponseRedirect("/listarClientesRetirados")
        
        elif modo == 'usuario':
            if request.user.is_superuser == False:
                messages.error(
                    request, 'No tienes permisos suficientes para borrar usuarios')
                return HttpResponseRedirect('/listarUsuarios')

            elif p == 1:
                messages.error(
                    request, 'No puedes eliminar al super-administrador.')
                return HttpResponseRedirect('/listarUsuarios')

            elif request.user.id == p:
                messages.error(
                    request, 'No puedes eliminar tu propio usuario.')
                return HttpResponseRedirect('/listarUsuarios')

            else:
                usuario = Usuario.objects.get(id=p)
                usuario.delete()
                messages.success(
                    request, 'Usuario de ID %s borrado exitosamente.' % p)
                return HttpResponseRedirect("/listarUsuarios")
# Fin de vista---------------------------------------------------------------------------------------


# Eliminar Clientes Retirados  ----------------------------------------------------------------------
class EliminarClienteRetirado(LoginRequiredMixin, View):
    login_url = '/login'
    redirect_field_name = None

    def get(self, request, modo, p):
        if modo == 'cliente_retirado':
            cliente_retirado = ClienteRetirado.objects.get(id=p)
            cliente = Cliente.objects.create(
                ip=cliente_retirado.ip,
                cedula=cliente_retirado.cedula,
                nombre=cliente_retirado.nombre,
                apellido=cliente_retirado.apellido,
                telefono_uno=cliente_retirado.telefono_uno,
                telefonos_dos=cliente_retirado.telefonos_dos,
                valor_instalacion=cliente_retirado.valor_instalacion,
                fecha_instalacion=cliente_retirado.fecha_instalacion,
                direccion=cliente_retirado.direccion,
                estado=cliente_retirado.estado,
                descripcion=cliente_retirado.descripcion,
                municipio=cliente_retirado.municipio,
                equipos=cliente_retirado.equipos,
                tipo_instalacion=cliente_retirado.tipo_instalacion,
                cap_megas=cliente_retirado.cap_megas,
                ap=cliente_retirado.ap
            )
            cliente_retirado.delete()
            messages.success(
                request, 'Cliente Retirado de ID %s movido a Clientes exitosamente.' % p)

            return redirect('listarClientes')

# Fin de vista---------------------------------------------------------------------------------------


# Crea una lista de la factura, 10 por pagina--------------------------------------------------------
class ListarFactura(LoginRequiredMixin,View):
    login_url = '/login'
    redirect_field_name = None

    def get(self, request, cliente_id):
        cliente = Cliente.objects.get(id=cliente_id)
        facturas = Factura.objects.filter(cliente=cliente)
        context = {'facturas': facturas}
        return render(request, 'factura/listarFacturas.html', context)

# Fin de vista---------------------------------------------------------------------------------------


# Crea una lista de la factura, 10 por pagina--------------------------------------------------------
class ListarFacturaRetirada(LoginRequiredMixin,View):
    login_url = '/login'
    redirect_field_name = None

    def get(self, request, id):
        cliente_retirado = ClienteRetirado.objects.get(id=id)
        factura_retirada = FacturaRetirada.objects.filter(cliente_retirado=cliente_retirado)
        contexto = {'factura_retirada': factura_retirada}
        return render(request, 'factura/listarFacturaRetirada.html', contexto)

# Fin de vista---------------------------------------------------------------------------------------


# Funcion que permite crear la factura de un cliente por su ID---------------------------------------

@login_required(login_url='/login')
def crear_factura(request, cliente_id):
    cliente = Cliente.objects.get(id=cliente_id)
    if request.method == 'POST':
        form = FacturaForm(request.POST)
        if form.is_valid():
            factura = form.save(commit=False)
            factura.cliente = cliente
            factura.save()
            return redirect('listarFactura', cliente.id)
    else:
        form = FacturaForm()
    return render(request, 'factura/emitirfactura.html', {'form': form})

# Fin de vista---------------------------------------------------------------------------------------


#Muestra los detalles individuales de una factura----------------------------------------------------
class VerFactura(LoginRequiredMixin,View):
    login_url = '/login'
    redirect_field_name = None
    def get(self, request, p):
        try:
            factura = Factura.objects.get(id=p)
        except Factura.DoesNotExist:
            return redirect('listarFactura')

        context = {'factura': factura}
        return render(request, 'factura/verFactura.html', context)
#Fin de vista----------------------------------------------------------------------------------------


#Muestra los detalles individuales de una factura----------------------------------------------------
class VerFacturaRetirada(LoginRequiredMixin,View):
    login_url = '/login'
    redirect_field_name = None
    def get(self, request, p):
        try:
            facturaRetirada = FacturaRetirada.objects.get(id=p)
        except FacturaRetirada.DoesNotExist:
            return redirect('listarClientesRetirados')

        context = {'facturaRetirada': facturaRetirada}
        return render(request, 'factura/verFacturaRetirada.html', context)
#Fin de vista----------------------------------------------------------------------------------------


#Genera la factura en PDF--------------------------------------------------------------------------
class GenerarFacturaPDF(LoginRequiredMixin,View):
    login_url = '/login'
    redirect_field_name = None

    def get(self, request, p):
        import io
        from reportlab.pdfgen import canvas
        from django.http import FileResponse
        from datetime import date  

        factura = Factura.objects.get(id=p)       
        data = { 
            'nombre':factura.cliente.nombre,
            'apellido': factura.cliente.apellido,
            'cedula': factura.cliente.cedula,
            'detalle': factura.detalle,
            'valor_pago': factura.valor_pago,
            'fecha': factura.fecha_pago,
            'valido_hasta': factura.fecha_vencimiento,
            'fecha':  date.today(),
        }
        nombre_factura = "factura_%s.pdf" % (factura.id)

        pdf = render_to_pdf('PDF/plantillaPDF.html', {'datos': [data]})

        response = HttpResponse(pdf,content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"' % nombre_factura

        return response  

#Fin de vista--------------------------------------------------------------------------------------


# #Genera la factura en PDF--------------------------------------------------------------------------
# class TodasFacturasPDF(LoginRequiredMixin, View):
#     login_url = '/login'
#     redirect_field_name = None
    
#     def get(self, request, cliente_id):
#         cliente = Cliente.objects.get(id=cliente_id)
#         facturas = Factura.objects.filter(cliente=cliente)
#         data = {'datos': facturas}

#         nombre_factura = f"facturas_{cliente_id}.pdf"

#         pdf = render_to_pdf('PDF/todasFacturas.html', {'datos': data, 'cliente': cliente})


#         response = HttpResponse(pdf, content_type='application/pdf')
#         response['Content-Disposition'] = f'attachment; filename="{nombre_factura}"'

#         return response


# #Fin de vista--------------------------------------------------------------------------------------


## Genera la factura en IMG--------------------------------------------------------------------------
# class GenerarIMG(LoginRequiredMixin,View):
#     login_url = '/login'
#     redirect_field_name = None
    
#     def get(self, request, cliente_id):
#         cliente = Cliente.objects.get(id=cliente_id)
#         facturas = Factura.objects.filter(cliente=cliente)

#         # Crea la imagen
#         image = Image.new('RGB', (600, 800), color='white')
#         draw = ImageDraw.Draw(image)
#         font = ImageFont.truetype('arial.ttf', size=20)

#         y_text = 50
#         for factura in facturas:
#             factura_text = f"Factura #{factura.id} - Fecha: {factura.fecha_pago} - Valor: {factura.valor_pago}"
#             draw.text((50, y_text), factura_text, font=font, fill=(0, 0, 0))
#             y_text += 50

#         # Guarda la imagen en un objeto BytesIO
#         img_byte_arr = BytesIO()
#         image.save(img_byte_arr, format='PNG')
#         img_byte_arr.seek(0)

#         # Devuelve la imagen como respuesta HTTP
#         response = HttpResponse(content_type="image/png")
#         response.write(img_byte_arr.read())
#         return response

## Fin de vista--------------------------------------------------------------------------------------


#Accede a los modulos del manual de usuario---------------------------------------------#
# class VerManualDeUsuario(LoginRequiredMixin, View):
#     login_url = '/inventario/login'
#     redirect_field_name = None

#     def get(self, request, pagina):
#         if pagina == 'inicio':
#             return render(request, 'manual/index.html') 

#         if pagina == 'producto':
#             return render(request, 'manual/producto.html') 

#         if pagina == 'proveedor':
#             return render(request, 'manual/proveedor.html') 

#         if pagina == 'pedido':
#             return render(request, 'manual/pedido.html') 

#         if pagina == 'clientes':
#             return render(request, 'manual/clientes.html') 

#         if pagina == 'factura':
#             return render(request, 'inventario/manual/factura.html') 

#         if pagina == 'usuarios':
#             return render(request, 'inventario/manual/usuarios.html')

#         if pagina == 'opciones':
#             return render(request, 'inventario/manual/opciones.html')



#Fin de vista--------------------------------------------------------------------------------


#Generar los reportes de los clientes Supendidos---------------------------------------------#
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

class ReporteSuspendidos(LoginRequiredMixin, TemplateView):
    login_url = '/login'
    redirect_field_name = None

    def get(self, request, *args, **kwargs):
        # Seleccionar los clientes en estado "Suspendidos"
        clientes_suspendidos = Cliente.objects.filter(estado='Suspendidos')

        # Crear el libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes Suspendidos"

        # Escribir los encabezados de la tabla
        ws['A1'] = 'Cédula'
        ws['B1'] = 'Ip'
        ws['C1'] = 'Nombre'
        ws['D1'] = 'Apellido'
        ws['E1'] = 'Telefono'
        ws['F1'] = 'Mensaje'
        ws['G1'] = 'WhatsApp'
        ws['A1'].font = Font(bold=True)  # Añadir formato negrita al encabezado
        ws['B1'].font = Font(bold=True)  # Añadir formato negrita al encabezado
        ws['C1'].font = Font(bold=True)  # Añadir formato negrita al encabezado
        ws['D1'].font = Font(bold=True)  # Añadir formato negrita al encabezado
        ws['E1'].font = Font(bold=True)  # Añadir formato negrita al encabezado
        ws['F1'].font = Font(bold=True)  # Añadir formato negrita al encabezado
        ws['G1'].font = Font(bold=True)  # Añadir formato negrita al encabezado

        # Escribir los datos de los clientes
        for i, cliente in enumerate(clientes_suspendidos, start=2):
            ws.cell(row=i, column=1, value=cliente.cedula)
            ws.cell(row=i, column=2, value=cliente.ip)
            ws.cell(row=i, column=3, value=cliente.nombre)
            ws.cell(row=i, column=4, value=cliente.apellido)
            ws.cell(row=i, column=5, value=cliente.telefono_uno)
 
            ws.cell(row=i, column=6, value='="⚠Estimado cliente {nombre} {apellido} queremos recordarle que su factura de internet  esta vencida, recuerde realizar su pago. Sí, ya realizo el pago POR FAVOR omita este mensaje."'.format(nombre=cliente.nombre, apellido=cliente.apellido))

            # Agregar la fórmula de WhatsApp para cada cliente
            cell = ws.cell(row=i, column=7)
            cell.value='=HYPERLINK("https://api.whatsapp.com/send?phone="&E{0}&"&text="&F{0}, "📲✔")'.format(i)
            cell.style = 'Hyperlink'  # Aplicar el estilo de hipervínculo al texto

        # Ajustar el ancho de las columnas
        for col in ws.columns:
            column = col[0].column_letter
            if column in ('A', 'B', 'D'):
                ws.column_dimensions[column].width = 15
            elif column == 'C':
                ws.column_dimensions[column].width = 30
            elif column == 'E':
                ws.column_dimensions[column].width = 20
            elif column == 'F':
                ws.column_dimensions[column].width = 25

        # Guardar el archivo y enviarlo como respuesta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=clientes_suspendidos.xlsx'
        wb.save(response)
        return response

    
#Fin de vista--------------------------------------------------------------------------------


#Generar los reportes de los clientes Para recoger equipo---------------------------------------------#
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

class ReporteREquipo(LoginRequiredMixin, TemplateView):
    login_url = '/login'
    redirect_field_name = None

    def get(self, request, *args, **kwargs):
        # Seleccionar los clientes en estado "Suspendidos"
        clientes_Requipos = Cliente.objects.filter(estado='REquipos')

        # Crear el libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes Para Retiro o en Rojo"

        # Escribir los encabezados de la tabla
        ws['A1'] = 'Cédula'
        ws['B1'] = 'Ip'
        ws['C1'] = 'Nombre'
        ws['D1'] = 'Apellido'
        ws['E1'] = 'Telefono'
        ws['F1'] = 'Mensaje'
        ws['G1'] = 'WhatsApp'
        ws['A1'].font = Font(bold=True)  # Añadir formato negrita al encabezado
        ws['B1'].font = Font(bold=True)  # Añadir formato negrita al encabezado
        ws['C1'].font = Font(bold=True)  # Añadir formato negrita al encabezado
        ws['D1'].font = Font(bold=True)  # Añadir formato negrita al encabezado
        ws['E1'].font = Font(bold=True)  # Añadir formato negrita al encabezado
        ws['F1'].font = Font(bold=True)  # Añadir formato negrita al encabezado
        ws['G1'].font = Font(bold=True)  # Añadir formato negrita al encabezado

        # Escribir los datos de los clientes
        for i, cliente in enumerate(clientes_Requipos, start=2):
            ws.cell(row=i, column=1, value=cliente.cedula)
            ws.cell(row=i, column=2, value=cliente.ip)
            ws.cell(row=i, column=3, value=cliente.nombre)
            ws.cell(row=i, column=4, value=cliente.apellido)
            ws.cell(row=i, column=5, value=cliente.telefono_uno)
 
            ws.cell(row=i, column=6, value='="⚠Estimado cliente {nombre} {apellido} queremos infórmale que, debido a que lleva más de un mes en mora, la empresa enviará a recoger los equipos y así se dará por terminado el contrato."'.format(nombre=cliente.nombre, apellido=cliente.apellido))

            # Agregar la fórmula de WhatsApp para cada cliente
            cell = ws.cell(row=i, column=7)
            cell.value='=HYPERLINK("https://api.whatsapp.com/send?phone="&E{0}&"&text="&F{0}, "📲✔")'.format(i)
            cell.style = 'Hyperlink'  # Aplicar el estilo de hipervínculo al texto

        # Ajustar el ancho de las columnas
        for col in ws.columns:
            column = col[0].column_letter
            if column in ('A', 'B', 'D'):
                ws.column_dimensions[column].width = 15
            elif column == 'C':
                ws.column_dimensions[column].width = 30
            elif column == 'E':
                ws.column_dimensions[column].width = 20
            elif column == 'F':
                ws.column_dimensions[column].width = 25

        # Guardar el archivo y enviarlo como respuesta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=clientes_recoequipo.xlsx'
        wb.save(response)
        return response

    
#Fin de vista--------------------------------------------------------------------------------


#Generar los reportes de los clientes Para recoger equipo---------------------------------------------#
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

class ReporteAp(LoginRequiredMixin, TemplateView):
    login_url = '/login'
    redirect_field_name = None

    def get(self, request, *args, **kwargs):
        ap_filtro = request.GET.get('ap', '').lower()  # Convertir a minúsculas y asignar un valor predeterminado en caso de que no se haya seleccionado ningún AP
        # Seleccionar los clientes filtrados
        cliente_filtro = Cliente.objects.filter(ap__nombreAp=ap_filtro)

        # Crear el libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Fallas"

        # Escribir los encabezados de la tabla
        ws['A1'] = 'Cédula'
        ws['B1'] = 'Ip'
        ws['C1'] = 'Nombre'
        ws['D1'] = 'Apellido'
        ws['E1'] = 'Telefono'
        ws['F1'] = 'Mensaje'
        ws['G1'] = 'AP'
        ws['H1'] = 'WhatsApp'
        ws['A1'].font = Font(bold=True)  # Añadir formato negrita al encabezado
        ws['B1'].font = Font(bold=True)  # Añadir formato negrita al encabezado
        ws['C1'].font = Font(bold=True)  # Añadir formato negrita al encabezado
        ws['D1'].font = Font(bold=True)  # Añadir formato negrita al encabezado
        ws['E1'].font = Font(bold=True)  # Añadir formato negrita al encabezado
        ws['F1'].font = Font(bold=True)  # Añadir formato negrita al encabezado
        ws['G1'].font = Font(bold=True)  # Añadir formato negrita al encabezado
        ws['h1'].font = Font(bold=True)  # Añadir formato negrita al encabezado

        # Escribir los datos de los clientes
        for i, cliente in enumerate(cliente_filtro, start=2):
            ws.cell(row=i, column=1, value=cliente.cedula)
            ws.cell(row=i, column=2, value=cliente.ip)
            ws.cell(row=i, column=3, value=cliente.nombre)
            ws.cell(row=i, column=4, value=cliente.apellido)
            ws.cell(row=i, column=5, value=cliente.telefono_uno)
 
            ws.cell(row=i, column=6, value='="⚠Estimado cliente {nombre} {apellido} queremos infórmale que estamos presentando fallas en su servicio de internet, estamos trabajando para solucionarlo."'.format(nombre=cliente.nombre, apellido=cliente.apellido))
            ws.cell(row=i, column=7, value=cliente.ap.nombreAp)
            # Agregar la fórmula de WhatsApp para cada cliente
            cell = ws.cell(row=i, column=8)
            cell.value='=HYPERLINK("https://api.whatsapp.com/send?phone="&E{0}&"&text="&F{0}, "📲✔")'.format(i)
            cell.style = 'Hyperlink'  # Aplicar el estilo de hipervínculo al texto

        # Ajustar el ancho de las columnas
        for col in ws.columns:
            column = col[0].column_letter
            if column in ('A', 'B', 'D'):
                ws.column_dimensions[column].width = 15
            elif column == 'C':
                ws.column_dimensions[column].width = 30
            elif column == 'E':
                ws.column_dimensions[column].width = 20
            elif column == 'F':
                ws.column_dimensions[column].width = 25

        # Guardar el archivo y enviarlo como respuesta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=reporte_fallas.xlsx'
        wb.save(response)
        return response

    
#Fin de vista--------------------------------------------------------------------------------


#Listar Suspendidos--------------------------------------------------------------------------------
class ListarSuspendidos(LoginRequiredMixin,TemplateView):
    login_url = '/login'
    redirect_field_name = None

    template_name = 'Reportes/reporteSuspendidos.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['clientes_suspendidos'] = Cliente.objects.filter(estado='Suspendidos')
        return context
#Fin de vista--------------------------------------------------------------------------------


#Listar Clientes Recoger Equipos--------------------------------------------------------------------------------
class ListarREquipos(LoginRequiredMixin,TemplateView):
    login_url = '/login'
    redirect_field_name = None

    template_name = 'Reportes/reporteREquipo.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['clientes_REquipos'] = Cliente.objects.filter(estado='REquipos')
        return context
#Fin de vista--------------------------------------------------------------------------------


#Listar Clientes Recoger Equipos--------------------------------------------------------------------------------
class ListarClientesAp(LoginRequiredMixin, TemplateView):
    login_url = '/login'
    redirect_field_name = None
    template_name = 'Reportes/reportesAp.html'

    def get(self, request):
        aps = Ap.objects.all()
        ap_filtro = request.GET.get('ap', '').lower()  # Convertir a minúsculas y asignar un valor predeterminado en caso de que no se haya seleccionado ningún AP
        if ap_filtro:
            clientes = Cliente.objects.filter(ap__nombreAp=ap_filtro)

        else:
            clientes = Cliente.objects.all()
        contexto = {'clientes': clientes, 'aps': aps}
        return render(request, 'Reportes/reportesAp.html', contexto)


#Fin de vista--------------------------------------------------------------------------------
